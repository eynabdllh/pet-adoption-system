from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
import openpyxl
from .forms import AdoptionForm
from pet_listing.models import Pet
from schedule_form.models import Schedule
from django.contrib.auth.decorators import login_required
from login_register.decorators import admin_required, adopter_required
from login_register.models import User
from profile_management.models import Profile  
from django.utils import timezone
from .models import Adoption
from django.core.paginator import Paginator
from django.db.models import Q

from notifications.models import Notification

@login_required
@adopter_required
def adopt_form(request, pet_id):
    user_id = request.session.get('user_id')
    pet = get_object_or_404(Pet, id=pet_id)

    if not user_id:
        return redirect('login')

    user = User.objects.get(id=user_id)
    profile, created = Profile.objects.get_or_create(user=user)  

    if request.method == 'POST':
        form = AdoptionForm(request.POST, user=user)
        if form.is_valid():
            request.session['adoption_form_data'] = {
                'adopter_id': profile.id,
                'pet_id': pet.id,
                'first_name': form.cleaned_data.get('first_name', user.first_name),
                'last_name': form.cleaned_data.get('last_name', user.last_name),
                'age': form.cleaned_data['age'],
                'address': form.cleaned_data['address'],
                'contact_number': form.cleaned_data['contact_number'],
                'email': form.cleaned_data.get('email', user.email),
                'date': form.cleaned_data['date'].isoformat(),
            }

            if not profile.age:
                profile.age = form.cleaned_data['age']
            if not profile.address:
                profile.address = form.cleaned_data['address']
            if not profile.phone_number:
                profile.phone_number = form.cleaned_data['contact_number']
            profile.save()

            return redirect('schedule', pet_id=pet.id)
    else:
        initial_data = {
            'adopter_id': user.id,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'age': profile.age if profile.age else '',
            'address': profile.address if profile.address else '',
            'contact_number': profile.phone_number if profile.phone_number else '',
        }

        form = AdoptionForm(initial=initial_data, user=user)

    today = timezone.localdate()

    has_notification = Notification.user_has_unread_notifs(user=request.session.get('user_id'))

    return render(request, 'adopt_form.html', {
        'form': form,
        'pet': pet,
        'today': today.isoformat(),
        'user': user,
        'has_notification': has_notification,
    })

@login_required
@admin_required
def adoption_management(request):
    status = request.GET.get('status', 'requested')
    pet_type = request.GET.get('pet_type', '')
    query = request.GET.get('q', '')
    sort_by_name = request.GET.get('sort_by_name', '')
    sort_by_id = request.GET.get('sort_by_id', '') 

    reset_filter = request.GET.get('reset_filter', False)
    reset_sort = request.GET.get('reset_sort', False)

    if reset_filter:
        pet_type = ''
    if reset_sort:
        sort_by_id = ''
        sort_by_name = ''

    if request.method == 'POST':
        action = request.POST.get('action')
        pet_id = request.POST.get('pet_id')

        if action == 'add_to_list' and pet_id:
            try:
                pet = Pet.objects.get(id=pet_id, is_rejected=True)
                pet.adoption_set.all().delete()
                Schedule.objects.filter(pet=pet).delete()

                pet.is_rejected = False
                pet.is_adopted = False
                pet.is_requested = False
                pet.is_available = True
                pet.save()

                messages.success(request, f"Pet {pet.name} is now available for adoption")
            except Pet.DoesNotExist:
                messages.error(request, "Pet not found or already updated.")

    if status == 'approved':
        pets = Pet.objects.filter(is_approved=True).prefetch_related('schedule_set')
    elif status == 'rejected':
        pets = Pet.objects.filter(is_rejected=True).prefetch_related('schedule_set')
    else:
        pets = Pet.objects.filter(is_requested=True).prefetch_related('schedule_set')

    # filter 
    if pet_type:
        pets = pets.filter(pet_type=pet_type)

    if query:
        pets = pets.filter(name__icontains=query)

    #sort
    if sort_by_id == 'asc':
        pets = pets.order_by('id')
    elif sort_by_id == 'desc':
        pets = pets.order_by('-id')

    if sort_by_name:
        if sort_by_name == 'asc':
            pets = pets.order_by('name')
        elif sort_by_name == 'desc':
            pets = pets.order_by('-name')

    has_notification = Notification.user_has_unread_notifs(user=request.session.get('user_id'))

    return render(request, 'adoption_management.html', {
        'pets': pets,
        'status': status,
        'sort_by_id': sort_by_id,
        'sort_by_name': sort_by_name,
        'pet_type': pet_type,
        'query': query,
        'has_notification': has_notification
    })

@login_required
@admin_required
def review_form(request, pet_id):
    adoption = get_object_or_404(Adoption.objects.select_related('adopter__user'), pet__id=pet_id)
    profile = adoption.adopter
    source = request.GET.get('source')
 
    if request.method == 'POST':
        status = request.POST.get('status')
        reason = request.POST.get('reason')
        pet = adoption.pet
        adopter = adoption.adopter.user
 
        try:
            if status == 'approved':
                if not pet.is_approved:
                    pet.is_requested = False
                    pet.is_approved = True
                    pet.is_upcoming = True
                    pet.save()
 
                    adoption.status = 'approved'
                    adoption.save()
 
                    messages.success(request, f"Pet {pet.name} has been approved for adoption.")
 
                    Notification.objects.create_for_adopter(
                        adopter=adopter,
                        title="Adoption Request Approved",
                        message=f"Your adoption request for {pet.name} has been approved! Please check your schedule for pickup details.",
                        pet=pet
                    )
 
            elif status == 'rejected':
                if not reason:
                    messages.error(request, "A reason is required for rejection.")
                else:
                    if not pet.is_rejected:
                        pet.is_requested = False
                        pet.is_rejected = True
                        pet.is_adopted = False
                        pet.save()
 
                        adoption.status = 'rejected'
                        adoption.reason_choices = reason
                        adoption.save()
 
                        messages.success(request, "The pet adoption request has been rejected.")
 
                        Notification.objects.create_for_adopter(
                            adopter=adopter,    
                            title="Adoption Request Rejected",
                            message=f"Your adoption request for {pet.name} has been rejected. Reason: {reason}",
                            pet=pet
                        )
 
            if source == 'admin_dashboard':
                return redirect('admin_dashboard')
            return redirect('adoption_management')
           
        except Exception as e:
            messages.error(request, f"An error occurred: {e}")

    has_notification = Notification.user_has_unread_notifs(user=request.session.get('user_id'))

    return render(request, 'review_form.html', {'adoption': adoption, 'profile': profile, 'has_notification':has_notification})


@login_required
@admin_required
def admin_pickup(request):
    if request.method == 'POST':
        pet_id = request.POST.get('pet_id')
        action = request.POST.get('action')

        if not pet_id:
            messages.error(request, "Pet ID is required.")
        else:
            try:
                pet = Pet.objects.get(id=pet_id) 
            except Pet.DoesNotExist:
                messages.error(request, "Pet not found.")
                pet = None  

            if pet and action == 'mark_completed': 
                try:
                    adoption = pet.adoption_set.first()
                    adopter = adoption.adopter.user if adoption else None
                    
                    pet.is_approved = False
                    pet.is_upcoming = False
                    pet.is_adopted = True
                    pet.save()
                    messages.success(request, f"Pet {pet.name} marked as completed.")

                    if adopter:
                        Notification.objects.create_for_adopter(
                            adopter=adopter,
                            title="Adoption Completed",
                            message=f"Congratulations! The adoption of {pet.name} has been completed successfully.",
                            pet=pet
                        )

                except Exception as e:
                    messages.error(request, f"An error occurred while marking as completed: {e}")

            if pet and action == 'mark_failed': 
                try:
                    adoption = pet.adoption_set.first()
                    adopter = adoption.adopter.user if adoption else None
                    
                    pet.is_approved = False
                    pet.is_upcoming = False
                    pet.is_adopted = False
                    pet.is_cancelled = True
                    pet.save()

                    messages.success(request, f"Pet {pet.name} was not picked up on time.")

                    if adopter:
                        Notification.objects.create_for_adopter(
                            adopter=adopter,
                            title="Adoption Cancelled - No Show",
                            message=f"Your adoption of {pet.name} has been cancelled as you didn't show up for pickup.",
                            pet=pet
                        )
                except Exception as e:
                    messages.error(request, f"An error occurred while marking as failed: {e}")

            if pet and action == 'add_to_list':  
                try:
                    pet.adoption_set.all().delete()  
                    Schedule.objects.filter(pet=pet).delete()  
                    pet.is_rejected = False
                    pet.is_adopted = False
                    pet.is_requested = False
                    pet.is_cancelled = False
                    pet.is_available = True
                    pet.save()

                    messages.success(request, f"Pet {pet.name} is now available for adoption.")
                    
                except Exception as e:
                    messages.error(request, f"An error occurred while adding to the list: {e}")

    status = request.GET.get('status', 'upcoming')
    pet_type = request.GET.get('pet_type', '')
    query = request.GET.get('q', '')
    sort_by_name = request.GET.get('sort_by_name', '')
    sort_by_id = request.GET.get('sort_by_id', '') 

    reset_filter = request.GET.get('reset_filter', False)
    reset_sort = request.GET.get('reset_sort', False)
    if reset_filter:
        pet_type = ''
    if reset_sort:
        sort_by_id = ''
        sort_by_name = ''

    if status == 'completed':
        pets = Pet.objects.filter(is_adopted=True, is_upcoming=False, is_approved=False).prefetch_related('schedule_set')
    elif status == 'cancelled':
        pets = Pet.objects.filter(is_cancelled=True).prefetch_related('schedule_set', 'adoption_set')
    else:
        pets = Pet.objects.filter(is_upcoming=True, is_approved=True).prefetch_related('schedule_set')

    if pet_type:
        pets = pets.filter(pet_type=pet_type)

    if query:
        pets = pets.filter(name__icontains=query)

    if sort_by_id == 'asc':
        pets = pets.order_by('id')
    elif sort_by_id == 'desc':
        pets = pets.order_by('-id')

    if sort_by_name:
        if sort_by_name == 'asc':
            pets = pets.order_by('name')
        elif sort_by_name == 'desc':
            pets = pets.order_by('-name')

    if status == 'cancelled':
        for pet in pets:
            schedules = pet.schedule_set.filter(cancelled=True)
            if schedules.exists():
                # to fetch the reason
                pet.cancellation_reason = schedules.first().get_reason_choices_display()
            else:
                # default if no reason is found
                pet.cancellation_reason = "Pet was not Picked-Up"

    has_notification = Notification.user_has_unread_notifs(user=request.session.get('user_id'))

    return render(request, 'admin_pickup.html', {
        'pets': pets,
        'status': status,
        'sort_by_id': sort_by_id,  
        'sort_by_name': sort_by_name, 
        'pet_type': pet_type,
        'query': query,
        'has_notification': has_notification,
    })


@login_required
@admin_required
def export_adoption_to_excel(request):
    pet_type = request.GET.get('pet_type', '')
    sort_by = request.GET.get('sort_by', '')
    status_filter = request.GET.get('status', '')  

    pets = Pet.objects.all()
    if pet_type:
        pets = pets.filter(pet_type__iexact=pet_type)
    
    if status_filter == 'pending':
        pets = pets.filter(is_requested=True)
    elif status_filter == 'approved':
        pets = pets.filter(is_approved=True)
    elif status_filter == 'rejected':
        pets = pets.filter(is_rejected=True)
    
    pets = pets.order_by(sort_by or 'id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pet Adoption List"

    headers = ['ID', 'Pet Name', 'Species', 'Adopter', 'Request Date', 'Status']
    if status_filter == 'rejected':
        headers.append('Reason')
    ws.append(headers)

    for pet in pets:
        adoption_data = pet.adoption_set.first()
        if adoption_data:
            adopter_name = f"{adoption_data.first_name} {adoption_data.last_name}"
            request_date = adoption_data.date.strftime('%Y-%m-%d') if adoption_data.date else 'N/A'
        else:
            adopter_name = 'N/A'
            request_date = 'N/A'

        if pet.is_requested:
            status = 'Pending'
        elif pet.is_approved:
            status = 'Approved'
        elif pet.is_rejected:
            status = 'Rejected'
        else:
            status = 'Unknown'

        row = [pet.id, pet.name, pet.pet_type, adopter_name, request_date, status]
        if status_filter == 'rejected':
            reason = adoption_data.get_reason_choices_display() if adoption_data else 'N/A'
            row.append(reason)
        ws.append(row)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="adoption_list_{status_filter}.xlsx"'
    wb.save(response)
    return response

@login_required
@admin_required
def export_pickup_to_excel(request):
    pet_type = request.GET.get('pet_type', '')
    status_filter = request.GET.get('status', '')
    sort_by = request.GET.get('sort_by', 'id')  

    valid_sort_fields = [
        'id', 'name', 'pet_type', 'age', 'adoption_fee', 'time_in_shelter',
        'gender', 'is_adopted', 'is_available', 'is_cancelled', 'is_requested'
    ]
    if sort_by not in valid_sort_fields:
        sort_by = 'id' 

    pets = Pet.objects.all()
    if pet_type:
        pets = pets.filter(pet_type__iexact=pet_type)
    
    if status_filter == 'upcoming':
        pets = pets.filter(is_upcoming=True, is_approved=True)
    elif status_filter == 'completed':
        pets = pets.filter(is_adopted=True, is_upcoming=False, is_approved=False)
    elif status_filter == 'cancelled':
        pets = pets.filter(is_cancelled=True)
    else:
        pets = pets.filter(is_upcoming=True) 

    pets = pets.order_by(sort_by)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pet List"

    if status_filter == 'cancelled':
        ws.append(['ID', 'Pet Name', 'Species', 'Adopter', 'Adoption Date', 'Status', 'Cancellation Reason'])
    else:
        ws.append(['ID', 'Pet Name', 'Species', 'Adopter', 'Adoption Date', 'Status'])

    for pet in pets:
        adoption_data = pet.adoption_set.first() 
        if adoption_data:
            adopter_name = f"{adoption_data.first_name} {adoption_data.last_name}"
            adoption_date = adoption_data.date
        else:
            adopter_name = 'N/A'
            adoption_date = 'N/A'

        if pet.is_upcoming and pet.is_approved:
            status = 'Upcoming'
        elif pet.is_adopted:
            status = 'Completed'
        elif pet.is_cancelled:
            status = 'Cancelled'
        else:
            status = 'Unknown'

        cancellation_reason = 'Pet was not Picked-Up'
        if pet.is_cancelled:
            schedules = pet.schedule_set.filter(cancelled=True)
            if schedules.exists():
                cancellation_reason = schedules.first().get_reason_choices_display()

        if status_filter == 'cancelled':
            ws.append([
                pet.id,
                pet.name,
                pet.pet_type.capitalize(),
                adopter_name,
                adoption_date,
                status,
                cancellation_reason
            ])
        else:
            ws.append([
                pet.id,
                pet.name,
                pet.pet_type.capitalize(),
                adopter_name,
                adoption_date,
                status
            ])

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{status_filter}_pickup_list.xlsx"'
    wb.save(response)
    return response
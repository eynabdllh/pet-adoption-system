import json
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
import openpyxl
from .forms import AdoptionForm
from pet_listing.models import Pet
from schedule_form.models import Schedule
from django.contrib.auth.decorators import login_required
from login_register.decorators import admin_required, adopter_required
from login_register.models import User
from profile_management.models import Profile  
from django.utils import timezone
from .models import Adoption

@login_required
@adopter_required
def adopt_form(request, pet_id):
    user_id = request.session.get('user_id')
    pet = get_object_or_404(Pet, id=pet_id)

    if not user_id:
        return redirect('login')

    user = User.objects.get(id=user_id)
    profile, created = Profile.objects.get_or_create(user=user)  

    if request.method == 'POST':
        form = AdoptionForm(request.POST, user=user)
        if form.is_valid():
            age = form.cleaned_data['age']
            address = form.cleaned_data['address']
            contact_number = form.cleaned_data['contact_number']

            if not profile.age:
                profile.age = age
            if not profile.address:
                profile.address = address
            if not profile.phone_number:
                profile.phone_number = contact_number
            profile.save()  

            adoption = Adoption.objects.create(
                adopter=profile,  
                pet=pet,
                first_name=form.cleaned_data.get('first_name', user.first_name),
                last_name=form.cleaned_data.get('last_name', user.last_name),
                age=age,
                address=address,
                contact_number=contact_number,
                email=form.cleaned_data.get('email', user.email),
                date=form.cleaned_data['date'],
            )

            pet.is_requested = True
            pet.is_available = False
            pet.save()

            request.session['adoption_form_data'] = {
                'adopter_id': profile.id,
                'pet_id': pet.id,
                'first_name': adoption.first_name,
                'last_name': adoption.last_name,
                'age': adoption.age,
                'address': adoption.address,
                'contact_number': adoption.contact_number,
                'email': adoption.email,
                'date': adoption.date.isoformat(),
            }

            return redirect('schedule', pet_id=pet.id)
    else:
        initial_data = {
            'adopter_id': user.id,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'age': profile.age if profile.age else '',
            'address': profile.address if profile.address else '',
            'contact_number': profile.phone_number if profile.phone_number else '',
        }

        form = AdoptionForm(initial=initial_data, user=user)

    today = timezone.localdate()
    return render(request, 'adopt_form.html', {
        'form': form,
        'pet': pet,
        'today': today.isoformat(),
        'user': user,
    })

@login_required
@admin_required
def adoption_management(request):
    status = request.GET.get('status', 'requested')  

    if request.method == 'POST':
        action = request.POST.get('action')
        pet_id = request.POST.get('pet_id')

        if action == 'add_to_list' and pet_id:
            try:
                pet = Pet.objects.get(id=pet_id, is_rejected=True)  
                
                pet.adoption_set.all().delete() 
                Schedule.objects.filter(pet=pet).delete()  
                
                pet.is_rejected = False
                pet.is_adopted = False
                pet.is_requested = False
                pet.is_available = True
                pet.save()
                
                messages.success(request, f"Pet {pet.name} is now available for adoption, and related records were deleted.")
            except Pet.DoesNotExist:
                messages.error(request, "Pet not found or already updated.")

    if status == 'approved':
        pets = Pet.objects.filter(is_approved=True).prefetch_related('adoption_set')
    elif status == 'rejected':
        pets = Pet.objects.filter(is_rejected=True).prefetch_related('adoption_set')
    else: 
        pets = Pet.objects.filter(is_requested=True).prefetch_related('adoption_set')

    return render(request, 'adoption_management.html', {
        'pets': pets,
        'status': status,  
    })

@login_required
@admin_required
def review_form(request, pet_id):
    adoption = get_object_or_404(Adoption.objects.select_related('adopter__user'), pet__id=pet_id)
    profile = adoption.adopter

    if request.method == 'POST':
        try:
            # Parse the incoming JSON request data
            data = json.loads(request.body)
            status = data.get('status')
            reason = data.get('reason')

            # Validate status - must be either 'approved' or 'rejected'
            if status not in ['approved', 'rejected']:
                messages.error(request, 'Invalid status provided.')
                return JsonResponse({'success': False, 'message': 'Invalid status provided.'})

            pet = adoption.pet

            if status == 'approved':
                # Approve the adoption request
                pet.is_requested = False
                pet.is_approved = True
                pet.is_upcoming = True
                pet.save()

                adoption.status = 'approved'
                adoption.save()

                messages.success(request, f"Pet {pet.name} has been approved for adoption.")
                return JsonResponse({'success': True, 'message': f"Pet {pet.name} has been approved for adoption."})

            elif status == 'rejected':
                # Ensure a reason is provided for rejection
                if not reason:
                    messages.error(request, "A reason is required for rejection.")
                    return JsonResponse({'success': False, 'message': "A reason is required for rejection."})

                # Reject the adoption request
                pet.is_requested = False
                pet.is_rejected = True
                pet.is_adopted = False
                pet.save()

                adoption.status = 'rejected'
                adoption.reason_choices = reason
                adoption.save()

                messages.success(request, "The pet adoption request has been rejected.")
                return JsonResponse({'success': True, 'message': "The pet adoption request has been rejected."})

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON format.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f"An error occurred: {str(e)}"})

    return render(request, 'review_form.html', {'adoption': adoption, 'profile': profile})

@login_required
@admin_required
def admin_pickup(request):
    if request.method == 'POST':
        pet_id = request.POST.get('pet_id')
        action = request.POST.get('action')

        if action == 'mark_completed' and pet_id:
            try:
                pet = Pet.objects.get(id=pet_id)
                pet.is_approved = False
                pet.is_upcoming = False
                pet.is_adopted = True
                pet.save()
                messages.success(request, f"Pet {pet.name} marked as completed.")
            except Pet.DoesNotExist:
                messages.error(request, "Pet not found.")
            except Exception as e:
                messages.error(request, f"An error occurred: {e}")

        if action == 'add_to_list' and pet_id:
            try:
                # Fetch the pet by ID
                pet = Pet.objects.get(id=pet_id)

                # Delete related adoption and schedule records
                pet.adoption_set.all().delete()  # Deletes all Adoption records linked to this pet
                Schedule.objects.filter(pet=pet).delete()  # Deletes all Schedule records linked to this pet

                # Update pet status
                pet.is_rejected = False
                pet.is_adopted = False
                pet.is_requested = False
                pet.is_cancelled = False  # Reset cancellation flag
                pet.is_available = True  # Make pet available for adoption
                pet.save()

                messages.success(request, f"Pet {pet.name} is now available for adoption, and related records were deleted.")
            except Pet.DoesNotExist:
                messages.error(request, "Pet not found.")
            except Exception as e:
                messages.error(request, f"An error occurred: {e}")

    # Retrieve pets based on status filter
    status = request.GET.get('status', 'upcoming')

    if status == 'completed':
        pets = Pet.objects.filter(is_adopted=True, is_upcoming=False, is_approved=False).prefetch_related('schedule_set')
    elif status == 'cancelled':
        pets = Pet.objects.filter(is_cancelled=True).prefetch_related('schedule_set', 'adoption_set')
    else: 
        pets = Pet.objects.filter(is_upcoming=True, is_approved=True).prefetch_related('schedule_set')

    # Add cancellation reasons to the context for cancelled pets
    for pet in pets:
        if status == 'cancelled':
            schedules = pet.schedule_set.filter(cancelled=True)
            for schedule in schedules:
                pet.cancellation_reason = schedule.get_reason_choices_display()

    return render(request, 'admin_pickup.html', {
        'pets': pets,
        'status': status,
    })


@login_required
@admin_required
def export_adoption_to_excel(request):
    pet_type = request.GET.get('pet_type', '')
    sort_by = request.GET.get('sort_by', '')
    status_filter = request.GET.get('status', '')  

    pets = Pet.objects.all()
    if pet_type:
        pets = pets.filter(pet_type__iexact=pet_type)
    
    if status_filter == 'pending':
        pets = pets.filter(is_requested=True)
    elif status_filter == 'approved':
        pets = pets.filter(is_approved=True)
    elif status_filter == 'rejected':
        pets = pets.filter(is_rejected=True)
    
    pets = pets.order_by(sort_by or 'id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pet Adoption List"

    headers = ['ID', 'Pet Name', 'Species', 'Adopter', 'Request Date', 'Status']
    if status_filter == 'rejected':
        headers.append('Reason')
    ws.append(headers)

    for pet in pets:
        adoption_data = pet.adoption_set.first()
        if adoption_data:
            adopter_name = f"{adoption_data.first_name} {adoption_data.last_name}"
            request_date = adoption_data.date.strftime('%Y-%m-%d') if adoption_data.date else 'N/A'
        else:
            adopter_name = 'N/A'
            request_date = 'N/A'

        if pet.is_requested:
            status = 'Pending'
        elif pet.is_approved:
            status = 'Approved'
        elif pet.is_rejected:
            status = 'Rejected'
        else:
            status = 'Unknown'

        row = [pet.id, pet.name, pet.pet_type, adopter_name, request_date, status]
        if status_filter == 'rejected':
            reason = adoption_data.get_reason_choices_display() if adoption_data else 'N/A'
            row.append(reason)
        ws.append(row)

    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="adoption_list_{status_filter}.xlsx"'
    wb.save(response)
    return response

@login_required
@admin_required
def export_pickup_to_excel(request):
    pet_type = request.GET.get('pet_type', '')
    status_filter = request.GET.get('status', '')
    sort_by = request.GET.get('sort_by', 'id')  

    valid_sort_fields = [
        'id', 'name', 'pet_type', 'age', 'adoption_fee', 'time_in_shelter',
        'gender', 'is_adopted', 'is_available', 'is_cancelled', 'is_requested'
    ]
    if sort_by not in valid_sort_fields:
        sort_by = 'id' 

    pets = Pet.objects.all()
    if pet_type:
        pets = pets.filter(pet_type__iexact=pet_type)
    
    if status_filter == 'upcoming':
        pets = pets.filter(is_upcoming=True, is_approved=True)
    elif status_filter == 'completed':
        pets = pets.filter(is_adopted=True, is_upcoming=False, is_approved=False)
    elif status_filter == 'cancelled':
        pets = pets.filter(is_cancelled=True)
    else:
        pets = pets.filter(is_upcoming=True) 

    pets = pets.order_by(sort_by)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pet List"

    if status_filter == 'cancelled':
        ws.append(['ID', 'Pet Name', 'Species', 'Adopter', 'Adoption Date', 'Status', 'Cancellation Reason'])
    else:
        ws.append(['ID', 'Pet Name', 'Species', 'Adopter', 'Adoption Date', 'Status'])

    for pet in pets:
        adoption_data = pet.adoption_set.first()  
        if adoption_data:
            adopter_name = f"{adoption_data.first_name} {adoption_data.last_name}"
            adoption_date = adoption_data.date
        else:
            adopter_name = 'N/A'
            adoption_date = 'N/A'

        if pet.is_upcoming and pet.is_approved:
            status = 'Upcoming'
        elif pet.is_adopted:
            status = 'Completed'
        elif pet.is_cancelled:
            status = 'Cancelled'
        else:
            status = 'Unknown'

        if status_filter == 'cancelled':
            ws.append([
                pet.id,
                pet.name,
                pet.pet_type.capitalize(),
                adopter_name,
                adoption_date,
                status,
                pet.cancellation_reason or 'N/A'
            ])
        else:
            ws.append([
                pet.id,
                pet.name,
                pet.pet_type.capitalize(),
                adopter_name,
                adoption_date,
                status
            ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{status_filter}_pickup_list.xlsx"'
    wb.save(response)
    return response
